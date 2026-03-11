"""
Genera cemaco_ga4_landing_report.xlsx con tres hojas:
  1. daily_detail    — detalle por landing page
  2. daily_summary   — resumen diario + validación VTEX
  3. dashboard       — gráficos y KPIs últimos 30 días
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.chart import LineChart, BarChart, Reference, Series
from openpyxl.chart.series import SeriesLabel
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.styles.differential import DifferentialStyle
from datetime import date, timedelta
import random

# ── Constantes de estilo ────────────────────────────────────────────────────
HEADER_BG   = "1F3864"   # azul oscuro
HEADER_FONT = "FFFFFF"   # blanco
ALT_ROW_BG  = "EBF0FA"   # azul muy claro (filas pares)

FILL_HEADER = PatternFill("solid", fgColor=HEADER_BG)
FONT_HEADER = Font(bold=True, color=HEADER_FONT, name="Calibri", size=11)
FONT_NORMAL = Font(name="Calibri", size=10)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

# Formatos de número
FMT_CURRENCY   = '#,##0.00'
FMT_PERCENT    = '0.00%'
FMT_INTEGER    = '#,##0'
FMT_DATE       = 'YYYY-MM-DD'

THIN_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
)


def style_header_row(ws, row_num, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill   = FILL_HEADER
        cell.font   = FONT_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER


def style_data_row(ws, row_num, num_cols, is_even=False):
    fill = PatternFill("solid", fgColor=ALT_ROW_BG) if is_even else None
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font      = FONT_NORMAL
        cell.alignment = ALIGN_LEFT
        cell.border    = THIN_BORDER
        if fill:
            cell.fill = fill


def auto_width(ws, extra=4):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                cell_len = len(str(cell.value)) if cell.value else 0
                max_len = max(max_len, cell_len)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + extra, 60)


# ── Hoja 1: daily_detail ────────────────────────────────────────────────────
DETAIL_HEADERS = [
    "date", "landingPage", "channel", "campaign", "device",
    "sessions", "engagedSessions", "engagementRate",
    "purchases", "revenue", "addToCarts",
    "conversionRate", "revenuePerSession",
]

DETAIL_FORMATS = {
    "date":              FMT_DATE,
    "sessions":          FMT_INTEGER,
    "engagedSessions":   FMT_INTEGER,
    "engagementRate":    FMT_PERCENT,
    "purchases":         FMT_INTEGER,
    "revenue":           FMT_CURRENCY,
    "addToCarts":        FMT_INTEGER,
    "conversionRate":    FMT_PERCENT,
    "revenuePerSession": FMT_CURRENCY,
}


def build_detail_sheet(ws):
    ws.title = "daily_detail"
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 28

    # Headers
    for col_idx, header in enumerate(DETAIL_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
    style_header_row(ws, 1, len(DETAIL_HEADERS))

    # Datos de muestra (últimos 5 días, 3 landings)
    sample_data = generate_sample_detail()
    for row_idx, row_data in enumerate(sample_data, start=2):
        is_even = (row_idx % 2 == 0)
        for col_idx, header in enumerate(DETAIL_HEADERS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(header))
            if header in DETAIL_FORMATS:
                cell.number_format = DETAIL_FORMATS[header]
        style_data_row(ws, row_idx, len(DETAIL_HEADERS), is_even)

    # Filtros automáticos
    ws.auto_filter.ref = f"A1:{get_column_letter(len(DETAIL_HEADERS))}1"
    auto_width(ws)


# ── Hoja 2: daily_summary ───────────────────────────────────────────────────
SUMMARY_HEADERS = [
    "date", "totalSessions", "totalPurchases", "totalRevenue",
    "avgConversionRate", "topLandingByRevenue", "topChannel",
    "vtexOrders", "ga4Purchases", "diferenciaPct", "statusValidacion",
    "organic_revenue", "paid_revenue", "email_revenue", "direct_revenue",
]

SUMMARY_FORMATS = {
    "date":              FMT_DATE,
    "totalSessions":     FMT_INTEGER,
    "totalPurchases":    FMT_INTEGER,
    "totalRevenue":      FMT_CURRENCY,
    "avgConversionRate": FMT_PERCENT,
    "vtexOrders":        FMT_INTEGER,
    "ga4Purchases":      FMT_INTEGER,
    "diferenciaPct":     '0.00"%"',
    "organic_revenue":   FMT_CURRENCY,
    "paid_revenue":      FMT_CURRENCY,
    "email_revenue":     FMT_CURRENCY,
    "direct_revenue":    FMT_CURRENCY,
}


def build_summary_sheet(ws):
    ws.title = "daily_summary"
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 28

    for col_idx, header in enumerate(SUMMARY_HEADERS, start=1):
        ws.cell(row=1, column=col_idx, value=header)
    style_header_row(ws, 1, len(SUMMARY_HEADERS))

    sample_data = generate_sample_summary()
    for row_idx, row_data in enumerate(sample_data, start=2):
        is_even = (row_idx % 2 == 0)
        for col_idx, header in enumerate(SUMMARY_HEADERS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(header))
            if header in SUMMARY_FORMATS:
                cell.number_format = SUMMARY_FORMATS[header]
        style_data_row(ws, row_idx, len(SUMMARY_HEADERS), is_even)

    ws.auto_filter.ref = f"A1:{get_column_letter(len(SUMMARY_HEADERS))}1"

    # ── Formato condicional columna statusValidacion (col K = 11)
    status_col = SUMMARY_HEADERS.index("statusValidacion") + 1
    status_col_letter = get_column_letter(status_col)
    data_range = f"{status_col_letter}2:{status_col_letter}{len(sample_data) + 10}"

    green_fill  = PatternFill(bgColor="C6EFCE")
    yellow_fill = PatternFill(bgColor="FFEB9C")
    red_fill    = PatternFill(bgColor="FFC7CE")
    green_font  = Font(color="276221", bold=True)
    yellow_font = Font(color="9C5700", bold=True)
    red_font    = Font(color="9C0006", bold=True)

    ws.conditional_formatting.add(
        data_range,
        CellIsRule(
            operator='containsText',
            formula=['"✅ OK"'],
            fill=green_fill,
            font=green_font,
        )
    )
    ws.conditional_formatting.add(
        data_range,
        FormulaRule(
            formula=[f'NOT(ISERROR(SEARCH("REVISAR",{status_col_letter}2)))'],
            fill=yellow_fill,
            font=yellow_font,
        )
    )
    ws.conditional_formatting.add(
        data_range,
        FormulaRule(
            formula=[f'NOT(ISERROR(SEARCH("ALERTA",{status_col_letter}2)))'],
            fill=red_fill,
            font=red_font,
        )
    )

    auto_width(ws)


# ── Hoja 3: dashboard ───────────────────────────────────────────────────────
def build_dashboard_sheet(ws, summary_ws):
    ws.title = "dashboard"

    # ── KPI boxes en la parte superior ──────────────────────────────────────
    kpi_labels = [
        "Revenue últimos 7 días",
        "Purchases últimos 7 días",
        "Conversión promedio",
        "Top landing de la semana",
    ]
    kpi_formulas = [
        "=IFERROR(SUMIF(daily_summary!A:A,\">=\"&(TODAY()-7),daily_summary!D:D),0)",
        "=IFERROR(SUMIF(daily_summary!A:A,\">=\"&(TODAY()-7),daily_summary!C:C),0)",
        "=IFERROR(AVERAGEIF(daily_summary!A:A,\">=\"&(TODAY()-7),daily_summary!E:E),0)",
        "=IFERROR(INDEX(daily_summary!F:F,MATCH(MAX(FILTER(daily_summary!D:D,daily_summary!A:A>=(TODAY()-7))),daily_summary!D:D,0)),\"(sin datos)\")",
    ]
    kpi_formats = [FMT_CURRENCY, FMT_INTEGER, FMT_PERCENT, "@"]

    kpi_cols = [1, 4, 7, 10]  # columnas de inicio de cada KPI box

    # Título del dashboard
    ws.merge_cells("A1:M1")
    title_cell = ws["A1"]
    title_cell.value = "📊 Cemaco GA4 — Dashboard de Landing Pages"
    title_cell.font      = Font(bold=True, size=16, color=HEADER_BG, name="Calibri")
    title_cell.alignment = ALIGN_CENTER
    ws.row_dimensions[1].height = 32

    for i, (label, formula, fmt, col) in enumerate(
        zip(kpi_labels, kpi_formulas, kpi_formats, kpi_cols)
    ):
        # Etiqueta
        label_cell = ws.cell(row=3, column=col, value=label)
        label_cell.font      = Font(bold=True, size=10, color=HEADER_FONT, name="Calibri")
        label_cell.alignment = ALIGN_CENTER
        label_cell.fill      = FILL_HEADER
        ws.merge_cells(
            start_row=3, start_column=col,
            end_row=3, end_column=col + 2
        )
        # Valor
        value_cell = ws.cell(row=4, column=col, value=formula)
        value_cell.number_format = fmt
        value_cell.font      = Font(bold=True, size=14, color=HEADER_BG, name="Calibri")
        value_cell.alignment = ALIGN_CENTER
        value_cell.border    = Border(
            left=Side(style='medium', color=HEADER_BG),
            right=Side(style='medium', color=HEADER_BG),
            bottom=Side(style='medium', color=HEADER_BG),
            top=Side(style='thin', color='CCCCCC'),
        )
        ws.merge_cells(
            start_row=4, start_column=col,
            end_row=5, end_column=col + 2
        )
        ws.row_dimensions[4].height = 30
        ws.row_dimensions[5].height = 30

    # ── Tabla resumen últimos 30 días ────────────────────────────────────────
    table_start_row = 8
    table_headers = ["Fecha", "Sesiones", "Purchases", "Revenue (Q)", "Conversión", "Estado VTEX"]
    table_col_map = [
        ("A", FMT_DATE), ("B", FMT_INTEGER), ("C", FMT_INTEGER),
        (FMT_CURRENCY, FMT_CURRENCY), (FMT_PERCENT, FMT_PERCENT), ("@", "@")
    ]

    ws.cell(row=table_start_row - 1, column=1, value="📋 Últimos 30 días").font = Font(
        bold=True, size=12, color=HEADER_BG
    )

    for col_idx, header in enumerate(table_headers, start=1):
        cell = ws.cell(row=table_start_row, column=col_idx, value=header)
    style_header_row(ws, table_start_row, len(table_headers))

    # Referencias dinámicas a daily_summary
    source_cols_summary = {
        "Fecha":        "A",
        "Sesiones":     "B",
        "Purchases":    "C",
        "Revenue (Q)":  "D",
        "Conversión":   "E",
        "Estado VTEX":  "K",
    }
    fmt_map = {
        "Fecha":        FMT_DATE,
        "Sesiones":     FMT_INTEGER,
        "Purchases":    FMT_INTEGER,
        "Revenue (Q)":  FMT_CURRENCY,
        "Conversión":   FMT_PERCENT,
        "Estado VTEX":  "@",
    }

    for dash_row in range(1, 31):
        excel_row = table_start_row + dash_row
        is_even = (dash_row % 2 == 0)
        for col_idx, header in enumerate(table_headers, start=1):
            src_col = source_cols_summary[header]
            # Referencia inversa: las últimas 30 filas de daily_summary
            formula = (
                f"=IFERROR(INDEX(daily_summary!{src_col}:{src_col},"
                f"COUNTA(daily_summary!A:A)-30+{dash_row}),\"\")"
            )
            cell = ws.cell(row=excel_row, column=col_idx, value=formula)
            cell.number_format = fmt_map[header]
        style_data_row(ws, excel_row, len(table_headers), is_even)

    # ── Gráfico de líneas: Revenue diario ────────────────────────────────────
    chart_row_start = table_start_row
    chart_row_end   = table_start_row + 30

    line_chart = LineChart()
    line_chart.title  = "Revenue Diario GA4 (Q)"
    line_chart.style  = 10
    line_chart.y_axis.title = "Revenue (Q)"
    line_chart.x_axis.title = "Fecha"
    line_chart.height = 14
    line_chart.width  = 24

    # Datos de revenue (columna D del dashboard = col 4)
    revenue_ref = Reference(ws, min_col=4, min_row=chart_row_start, max_row=chart_row_end)
    dates_ref   = Reference(ws, min_col=1, min_row=chart_row_start + 1, max_row=chart_row_end)
    line_chart.add_data(revenue_ref, titles_from_data=True)
    line_chart.set_categories(dates_ref)
    line_chart.series[0].graphicalProperties.line.solidFill = HEADER_BG
    line_chart.series[0].graphicalProperties.line.width     = 20000

    ws.add_chart(line_chart, "H8")

    # ── Gráfico de barras apiladas: Revenue por canal (últimos 7 días) ───────
    bar_chart = BarChart()
    bar_chart.type    = "col"
    bar_chart.grouping = "stacked"
    bar_chart.title   = "Revenue por Canal — Últimos 7 días"
    bar_chart.style   = 10
    bar_chart.y_axis.title = "Revenue (Q)"
    bar_chart.height  = 14
    bar_chart.width   = 24

    # Tabla auxiliar con datos de canal (últimos 7 días) desde daily_summary
    aux_start_col = 15
    aux_start_row = table_start_row
    channel_labels = ["Fecha", "Orgánico", "Paid", "Email", "Directo"]
    channel_src    = ["A",     "L",        "M",    "N",     "O"]

    for col_idx, label in enumerate(channel_labels, start=aux_start_col):
        cell = ws.cell(row=aux_start_row, column=col_idx, value=label)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = ALIGN_CENTER

    for dash_row in range(1, 8):
        excel_row = aux_start_row + dash_row
        for col_idx, src_col in enumerate(channel_src, start=aux_start_col):
            formula = (
                f"=IFERROR(INDEX(daily_summary!{src_col}:{src_col},"
                f"COUNTA(daily_summary!A:A)-7+{dash_row}),0)"
            )
            cell = ws.cell(row=excel_row, column=col_idx, value=formula)
            if src_col != "A":
                cell.number_format = FMT_CURRENCY

    colors = ["2E75B6", "ED7D31", "A9D18E", "FF0000"]
    for i, channel_name in enumerate(["Orgánico", "Paid", "Email", "Directo"]):
        col_num = aux_start_col + 1 + i
        data_ref = Reference(
            ws,
            min_col=col_num,
            min_row=aux_start_row,
            max_row=aux_start_row + 7
        )
        series = Series(data_ref, title_from_data=True)
        series.graphicalProperties.solidFill = colors[i]
        bar_chart.append(series)

    cats_ref = Reference(ws, min_col=aux_start_col, min_row=aux_start_row + 1, max_row=aux_start_row + 7)
    bar_chart.set_categories(cats_ref)

    ws.add_chart(bar_chart, "H26")

    auto_width(ws)


# ── Datos de muestra ─────────────────────────────────────────────────────────
SAMPLE_LANDINGS = [
    "/ofertas/verano-2026",
    "/categoria/pisos",
    "/producto/pintura-exterior",
    "/categoria/herramientas",
    "/home",
]
SAMPLE_CHANNELS = ["Organic Search", "Paid Search", "Direct", "Email", "Organic Social"]
SAMPLE_CAMPAIGNS = ["verano_2026", "(not set)", "dia_madre", "(not set)", "remarketing"]
SAMPLE_DEVICES = ["desktop", "mobile", "tablet"]


def generate_sample_detail():
    rows = []
    today = date.today()
    for days_ago in range(5, 0, -1):
        d = today - timedelta(days=days_ago)
        for landing, channel, campaign in zip(SAMPLE_LANDINGS, SAMPLE_CHANNELS, SAMPLE_CAMPAIGNS):
            sessions    = random.randint(200, 2000)
            purchases   = random.randint(5, int(sessions * 0.08))
            revenue     = round(purchases * random.uniform(150, 800), 2)
            engaged     = int(sessions * random.uniform(0.4, 0.75))
            eng_rate    = round(engaged / sessions, 4)
            add_carts   = int(purchases * random.uniform(1.5, 4.0))
            conv_rate   = round(purchases / sessions, 4)
            rev_per_ses = round(revenue / sessions, 2)
            rows.append({
                "date":             d,
                "landingPage":      landing,
                "channel":          channel,
                "campaign":         campaign,
                "device":           random.choice(SAMPLE_DEVICES),
                "sessions":         sessions,
                "engagedSessions":  engaged,
                "engagementRate":   eng_rate,
                "purchases":        purchases,
                "revenue":          revenue,
                "addToCarts":       add_carts,
                "conversionRate":   conv_rate,
                "revenuePerSession": rev_per_ses,
            })
    return rows


def generate_sample_summary():
    rows = []
    today = date.today()
    for days_ago in range(30, 0, -1):
        d = today - timedelta(days=days_ago)
        total_sessions  = random.randint(1500, 8000)
        total_purchases = random.randint(30, 200)
        total_revenue   = round(total_purchases * random.uniform(200, 700), 2)
        avg_conv        = round(total_purchases / total_sessions, 4)
        vtex_orders     = total_purchases + random.randint(-5, 15)
        dif_pct         = round(abs(vtex_orders - total_purchases) / max(vtex_orders, 1) * 100, 2)

        if dif_pct < 10:
            status = "✅ OK"
        elif dif_pct <= 20:
            status = "⚠️ REVISAR"
        else:
            status = "🚨 ALERTA"

        org_rev    = round(total_revenue * random.uniform(0.3, 0.5), 2)
        paid_rev   = round(total_revenue * random.uniform(0.2, 0.4), 2)
        email_rev  = round(total_revenue * random.uniform(0.05, 0.15), 2)
        direct_rev = round(total_revenue - org_rev - paid_rev - email_rev, 2)

        rows.append({
            "date":               d,
            "totalSessions":      total_sessions,
            "totalPurchases":     total_purchases,
            "totalRevenue":       total_revenue,
            "avgConversionRate":  avg_conv,
            "topLandingByRevenue": random.choice(SAMPLE_LANDINGS),
            "topChannel":         random.choice(["Organic Search", "Paid Search"]),
            "vtexOrders":         vtex_orders,
            "ga4Purchases":       total_purchases,
            "diferenciaPct":      dif_pct,
            "statusValidacion":   status,
            "organic_revenue":    org_rev,
            "paid_revenue":       paid_rev,
            "email_revenue":      email_rev,
            "direct_revenue":     direct_rev,
        })
    return rows


# ── Entrypoint ───────────────────────────────────────────────────────────────
def main():
    wb = openpyxl.Workbook()

    # Hoja 1
    ws_detail = wb.active
    build_detail_sheet(ws_detail)

    # Hoja 2
    ws_summary = wb.create_sheet()
    build_summary_sheet(ws_summary)

    # Hoja 3
    ws_dashboard = wb.create_sheet()
    build_dashboard_sheet(ws_dashboard, ws_summary)

    output_path = "cemaco_ga4_landing_report.xlsx"
    wb.save(output_path)
    print(f"✅ Archivo generado: {output_path}")


if __name__ == "__main__":
    main()
